#!/usr/bin/env python3
"""
Paper Trading Bot V4 - Live paper trading met Excel logging
------------------------------------------------------------
Draait de DualConfirmStrategy V4 OPTIMIZED op echte Kraken marktdata.
Geen echte orders - logt elke trade naar een Excel bestand.

V4 Optimalisaties (backtest: PF 78.29 | DD 1.9% | WR 83.3%):
  - Entry: Volume spike >2.0x + volume bar-to-bar confirm (V4)
  - Exit: Break-even stop na +3% winst
  - Exit: ATR trailing stop 2.0x
  - Exit: Time max 10 bars (40h) force close (V4: was 16)
  - Portfolio: 1x$2000 all-in + volume ranking
  - Smart cooldown: 8 bars na stop loss

Gebruik:
    python paper_trade.py              # Start paper trading (2 dagen)
    python paper_trade.py --hours 48   # Custom duur in uren
    python paper_trade.py --report     # Toon laatste rapport

Configuratie uit .env (STRATEGY=dual, CAPITAL_PER_TRADE, MAX_OPEN_POSITIONS)
"""
import os
import sys
import json
import time
import logging
import argparse
from datetime import datetime, timedelta, timezone
from pathlib import Path
from dotenv import load_dotenv

# Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# Project imports
from kraken_client import KrakenClient
from strategy import DualConfirmStrategy, DonchianBounceStrategy, MeanReversionStrategy, Position, Signal

# === SETUP ===
BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / '.env')

EXCEL_FILE = BASE_DIR / 'paper_trades.xlsx'  # Default, overridden by --tag
STATE_FILE = BASE_DIR / 'paper_state.json'   # Default, overridden by --tag
LOG_FILE = BASE_DIR / 'logs' / f"paper_{datetime.now().strftime('%Y%m%d_%H%M')}.log"

# Kraken fee (taker)
KRAKEN_FEE = 0.0026  # 0.26%


def setup_logging():
    """Configureer logging."""
    log_dir = BASE_DIR / 'logs'
    log_dir.mkdir(exist_ok=True)

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        handlers=[
            logging.FileHandler(LOG_FILE),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger('paper_trade')


# === STATE MANAGEMENT ===

def load_state() -> dict:
    """Laad paper trade state."""
    if STATE_FILE.exists():
        with open(STATE_FILE, 'r') as f:
            return json.load(f)
    return {
        'positions': {},
        'start_time': datetime.now(timezone.utc).isoformat(),
        'total_trades': 0,
        'total_pnl': 0.0,
        'wins': 0,
        'losses': 0,
        'checks': 0,
    }


def save_state(state: dict):
    """Sla state op."""
    with open(STATE_FILE, 'w') as f:
        json.dump(state, f, indent=2)


# === EXCEL MANAGEMENT ===

def create_excel():
    """Maak nieuw Excel bestand met headers en opmaak."""
    wb = Workbook()

    # === Sheet 1: Trades ===
    ws = wb.active
    ws.title = "Trades"

    headers = [
        'Nr', 'Datum/Tijd', 'Type', 'Coin', 'Entry Prijs', 'Exit Prijs',
        'Volume', 'Bedrag ($)', 'Fee ($)', 'P&L ($)', 'P&L (%)',
        'Exit Reden', 'RSI', 'Duur (uren)', 'Cumulatief P&L ($)'
    ]

    # Header styling
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    widths = [5, 18, 6, 12, 12, 12, 14, 12, 8, 10, 9, 25, 6, 10, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = 'A2'

    # === Sheet 2: Overzicht ===
    ws2 = wb.create_sheet("Overzicht")
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18

    overview_items = [
        ('PAPER TRADING RAPPORT', ''),
        ('', ''),
        ('Start', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('Einde', '(lopend)'),
        ('Strategie', 'Dual Confirm (DC+BB)'),
        ('', ''),
        ('RESULTATEN', ''),
        ('Totaal trades', 0),
        ('Winstgevend', 0),
        ('Verliesgevend', 0),
        ('Win rate', '0%'),
        ('', ''),
        ('Totaal P&L', '$0.00'),
        ('Gemiddelde P&L per trade', '$0.00'),
        ('Grootste winst', '$0.00'),
        ('Grootste verlies', '$0.00'),
        ('Profit Factor', 'N/A'),
        ('', ''),
        ('CONFIGURATIE V4', ''),
        ('Max posities', os.getenv('MAX_OPEN_POSITIONS', '1')),
        ('Capital per trade', f"${os.getenv('CAPITAL_PER_TRADE', '2000')}"),
        ('Kraken fee', '0.26% per side'),
        ('RSI DC/BB max', '40'),
        ('ATR stop mult', '2.0'),
        ('Cooldown', '4 bars (8 na stop)'),
        ('Max stop loss', '15%'),
        ('Volume spike filter', '>2.0x avg'),
        ('Break-even stop', 'Na +3% winst'),
        ('Time max', '10 bars / 40h (V4)'),
        ('Vol confirm', '>1.0x prev bar (V4)'),
        ('Volume min', '50% avg'),
    ]

    title_font = Font(bold=True, size=14, color='2F5496')
    section_font = Font(bold=True, size=11, color='2F5496')
    normal_font = Font(size=11)

    for row, (key, val) in enumerate(overview_items, 1):
        cell_a = ws2.cell(row=row, column=1, value=key)
        cell_b = ws2.cell(row=row, column=2, value=val)

        if row == 1:
            cell_a.font = title_font
        elif key in ('RESULTATEN', 'CONFIGURATIE', 'CONFIGURATIE V4'):
            cell_a.font = section_font
        else:
            cell_a.font = normal_font
            cell_b.font = normal_font

    # === Sheet 3: Open Posities ===
    ws3 = wb.create_sheet("Open Posities")
    pos_headers = ['Coin', 'Entry Prijs', 'Entry Tijd', 'Volume', 'Bedrag ($)',
                   'Stop Prijs', 'Target Prijs', 'RSI bij entry']

    for col, header in enumerate(pos_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    pos_widths = [12, 12, 18, 14, 12, 12, 12, 10]
    for i, w in enumerate(pos_widths, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    ws3.freeze_panes = 'A2'

    wb.save(EXCEL_FILE)
    return wb


def log_trade_to_excel(trade_data: dict, trade_nr: int, cumulative_pnl: float):
    """Log een afgesloten trade naar Excel."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Trades"]

    row = ws.max_row + 1
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    values = [
        trade_nr,
        trade_data.get('timestamp', ''),
        trade_data.get('type', ''),
        trade_data.get('pair', '').replace('/USD', ''),
        trade_data.get('entry_price', 0),
        trade_data.get('exit_price', 0),
        trade_data.get('volume', 0),
        trade_data.get('amount_usd', 0),
        trade_data.get('total_fee', 0),
        trade_data.get('pnl', 0),
        trade_data.get('pnl_pct', 0),
        trade_data.get('exit_reason', ''),
        trade_data.get('rsi', 0),
        trade_data.get('duration_hours', 0),
        cumulative_pnl,
    ]

    # Color based on P&L
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    red_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
    pnl = trade_data.get('pnl', 0)
    row_fill = green_fill if pnl >= 0 else red_fill

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal='center')

        # Number formatting
        if col in (5, 6):  # Prices
            cell.number_format = '#,##0.0000'
        elif col in (7,):  # Volume
            cell.number_format = '#,##0.00000000'
        elif col in (8, 9, 10, 15):  # USD amounts
            cell.number_format = '$#,##0.00'
        elif col == 11:  # P&L %
            cell.number_format = '0.00"%"'
        elif col == 14:  # Duration
            cell.number_format = '0.0'

    wb.save(EXCEL_FILE)


def log_buy_to_excel(trade_data: dict, trade_nr: int):
    """Log een BUY (open positie) naar Excel - als tussentijds overzicht."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Trades"]

    row = ws.max_row + 1
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    blue_fill = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')

    entry_fee = trade_data.get('amount_usd', 0) * KRAKEN_FEE
    values = [
        trade_nr,
        trade_data.get('timestamp', ''),
        'BUY',
        trade_data.get('pair', '').replace('/USD', ''),
        trade_data.get('entry_price', 0),
        '',  # No exit price yet
        trade_data.get('volume', 0),
        trade_data.get('amount_usd', 0),
        entry_fee,
        '',  # No P&L yet
        '',
        f"Stop: ${trade_data.get('stop_price', 0):.4f} | Target: ${trade_data.get('target_price', 0):.4f}",
        trade_data.get('rsi', 0),
        '',  # No duration yet
        '',  # No cumulative
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = thin_border
        cell.fill = blue_fill
        cell.alignment = Alignment(horizontal='center')

        if col in (5,) and isinstance(val, (int, float)):
            cell.number_format = '#,##0.0000'
        elif col in (7,) and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00000000'
        elif col in (8, 9) and isinstance(val, (int, float)):
            cell.number_format = '$#,##0.00'

    wb.save(EXCEL_FILE)


def update_open_positions_sheet(positions: dict):
    """Update de Open Posities sheet."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Open Posities"]

    # Clear old data (keep header)
    for row in range(2, ws.max_row + 1):
        for col in range(1, 9):
            ws.cell(row=row, column=col, value=None)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 2
    for pair, pos in positions.items():
        values = [
            pair.replace('/USD', ''),
            pos.get('entry_price', 0),
            datetime.fromtimestamp(pos.get('entry_time', 0)).strftime('%Y-%m-%d %H:%M'),
            pos.get('volume', 0),
            pos.get('entry_price', 0) * pos.get('volume', 0),
            pos.get('stop_price', 0),
            pos.get('target_price', 0),
            pos.get('rsi', 0),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
            if col in (2, 6, 7):
                cell.number_format = '#,##0.0000'
            elif col == 4:
                cell.number_format = '#,##0.00000000'
            elif col == 5:
                cell.number_format = '$#,##0.00'

        row += 1

    wb.save(EXCEL_FILE)


def update_overview_sheet(state: dict, positions: dict):
    """Update het overzicht sheet met huidige resultaten."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Overzicht"]

    total_trades = state.get('closed_trades', 0)
    wins = state.get('wins', 0)
    losses = state.get('losses', 0)
    total_pnl = state.get('total_pnl', 0.0)
    biggest_win = state.get('biggest_win', 0.0)
    biggest_loss = state.get('biggest_loss', 0.0)
    gross_wins = state.get('gross_wins', 0.0)
    gross_losses = abs(state.get('gross_losses', 0.0))

    win_rate = (wins / total_trades * 100) if total_trades > 0 else 0
    avg_pnl = (total_pnl / total_trades) if total_trades > 0 else 0
    pf = (gross_wins / gross_losses) if gross_losses > 0 else 'N/A'

    # Update values (row positions based on create_excel layout)
    ws.cell(row=4, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M') + ' (lopend)')
    ws.cell(row=8, column=2, value=total_trades)
    ws.cell(row=9, column=2, value=wins)
    ws.cell(row=10, column=2, value=losses)
    ws.cell(row=11, column=2, value=f'{win_rate:.1f}%')
    ws.cell(row=13, column=2, value=f'${total_pnl:.2f}')
    ws.cell(row=14, column=2, value=f'${avg_pnl:.2f}')
    ws.cell(row=15, column=2, value=f'${biggest_win:.2f}')
    ws.cell(row=16, column=2, value=f'${biggest_loss:.2f}')
    ws.cell(row=17, column=2, value=f'{pf:.2f}' if isinstance(pf, float) else pf)

    # Open positions info
    open_count = len(positions)
    if open_count > 0:
        row = 30
        ws.cell(row=row, column=1, value='OPEN POSITIES').font = Font(bold=True, size=11, color='2F5496')
        ws.cell(row=row, column=2, value=str(open_count))
        for pair, pos in positions.items():
            row += 1
            ws.cell(row=row, column=1, value=pair.replace('/USD', ''))
            ws.cell(row=row, column=2, value=f"${pos.get('entry_price', 0):.4f}")

    wb.save(EXCEL_FILE)


# === PAPER TRADING ENGINE ===

class PaperTrader:
    """Paper trading engine met Excel logging."""

    def __init__(self, duration_hours=48, capital=None, max_positions=None, tag=None):
        global EXCEL_FILE, STATE_FILE, LOG_FILE

        # Tag voor unieke bestanden per variant
        if tag:
            EXCEL_FILE = BASE_DIR / f'paper_trades_{tag}.xlsx'
            STATE_FILE = BASE_DIR / f'paper_state_{tag}.json'
            LOG_FILE = BASE_DIR / 'logs' / f"paper_{tag}_{datetime.now().strftime('%Y%m%d_%H%M')}.log"

        self.logger = setup_logging()
        self.duration_hours = duration_hours
        self.state = load_state()
        self.trade_counter = self.state.get('total_trades', 0)

        # Kraken client (alleen voor data, geen orders)
        api_key = os.getenv('KRAKEN_API_KEY', '')
        private_key = os.getenv('KRAKEN_PRIVATE_KEY', '')
        if not api_key or not private_key:
            self.logger.error("KRAKEN_API_KEY of KRAKEN_PRIVATE_KEY niet ingesteld")
            sys.exit(1)
        self.client = KrakenClient(api_key, private_key)

        # Configuratie - CLI overrides .env
        coins_str = os.getenv('COINS', '')
        self.coins = [c.strip() for c in coins_str.split(',') if c.strip()]

        self.capital_per_trade = capital if capital else float(os.getenv('CAPITAL_PER_TRADE', '1000'))
        self.max_positions = max_positions if max_positions else int(os.getenv('MAX_OPEN_POSITIONS', '2'))

        # Strategy V4 OPTIMIZED (altijd dual confirm voor paper trade)
        self.strategy = DualConfirmStrategy(
            rsi_dc_max=40, rsi_bb_max=40, rsi_sell=70,
            atr_stop_mult=2.0, cooldown_bars=4,
            cooldown_after_stop=8,
            max_stop_loss_pct=15.0,
            # V4 optimalisaties
            volume_spike_filter=True, volume_spike_mult=2.0,
            breakeven_stop=True, breakeven_trigger_pct=3.0,
            volume_min_pct=0.5,
            time_max_bars=10,
            vol_confirm=True, vol_confirm_mult=1.0,
        )

        # Per-coin strategy instances (voor cooldown tracking)
        self.strategies = {}

        # Excel
        if not EXCEL_FILE.exists():
            create_excel()

        self.logger.info(f"Paper Trader gestart")
        self.logger.info(f"  Duur: {duration_hours} uur")
        self.logger.info(f"  Coins: {len(self.coins)}")
        self.logger.info(f"  Capital/trade: ${self.capital_per_trade}")
        self.logger.info(f"  Max posities: {self.max_positions}")
        self.logger.info(f"  Fee: {KRAKEN_FEE*100:.2f}%")

    def get_strategy(self, pair: str) -> DualConfirmStrategy:
        """Haal strategy instance op per coin (voor cooldown tracking)."""
        if pair not in self.strategies:
            self.strategies[pair] = DualConfirmStrategy(
                rsi_dc_max=40, rsi_bb_max=40, rsi_sell=70,
                atr_stop_mult=2.0, cooldown_bars=4,
                cooldown_after_stop=8,
                max_stop_loss_pct=15.0,
                # V4 optimalisaties
                volume_spike_filter=True, volume_spike_mult=2.0,
                breakeven_stop=True, breakeven_trigger_pct=3.0,
                volume_min_pct=0.5,
                time_max_bars=10,
                vol_confirm=True, vol_confirm_mult=1.0,
            )
        return self.strategies[pair]

    def get_position(self, pair: str):
        """Haal positie op uit state."""
        pos_data = self.state.get('positions', {}).get(pair)
        if pos_data:
            return Position(
                pair=pos_data['pair'],
                entry_price=pos_data['entry_price'],
                volume=pos_data['volume'],
                stop_price=pos_data['stop_price'],
                highest_price=pos_data['highest_price'],
                entry_time=pos_data['entry_time'],
            )
        return None

    def check_all_coins(self):
        """Check alle coins voor signalen - 1 cycle."""
        self.state['checks'] = self.state.get('checks', 0) + 1
        check_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        self.logger.info(f"\n{'='*60}")
        self.logger.info(f"PAPER TRADE CHECK #{self.state['checks']} - {check_time}")
        self.logger.info(f"Open posities: {len(self.state.get('positions', {}))}/{self.max_positions}")
        self.logger.info(f"{'='*60}")

        signals_found = 0
        errors = 0

        for pair in self.coins:
            try:
                # Haal 4H candles op
                candles = self.client.get_ohlc(pair, interval=240)
                if not candles or len(candles) < 30:
                    continue

                position = self.get_position(pair)
                strategy = self.get_strategy(pair)
                signal = strategy.analyze(candles, position, pair)

                if signal.action == 'BUY':
                    # Check max positions
                    num_open = len(self.state.get('positions', {}))
                    if num_open >= self.max_positions:
                        self.logger.info(f"  SKIP {pair}: Max posities bereikt ({num_open}/{self.max_positions})")
                        continue

                    # Check niet al in positie
                    if pair in self.state.get('positions', {}):
                        continue

                    signals_found += 1
                    self._paper_buy(pair, signal, candles[-1])

                elif signal.action in ('SELL_TARGET', 'SELL_STOP'):
                    if position:
                        signals_found += 1
                        self._paper_sell(pair, signal, position)

                elif signal.action == 'HOLD' and position:
                    # Update highest price tracking
                    current_price = candles[-1]['close']
                    if current_price > position.highest_price:
                        self.state['positions'][pair]['highest_price'] = current_price
                    # Update stop price if strategy raised it
                    if position.stop_price > self.state['positions'][pair].get('stop_price', 0):
                        self.state['positions'][pair]['stop_price'] = position.stop_price

            except Exception as e:
                errors += 1
                if errors <= 5:  # Limit error logging
                    self.logger.warning(f"  Error {pair}: {e}")

        # Update Excel
        update_open_positions_sheet(self.state.get('positions', {}))
        update_overview_sheet(self.state, self.state.get('positions', {}))
        save_state(self.state)

        self.logger.info(f"\nCheck compleet: {signals_found} signalen, {errors} errors")
        self.logger.info(f"Open posities: {len(self.state.get('positions', {}))}/{self.max_positions}")
        self.logger.info(f"Totaal P&L: ${self.state.get('total_pnl', 0):.2f}")

    def _paper_buy(self, pair: str, signal: Signal, candle: dict):
        """Simuleer een buy."""
        volume = self.capital_per_trade / signal.price
        entry_fee = self.capital_per_trade * KRAKEN_FEE

        self.trade_counter += 1

        self.logger.info(f"  BUY #{self.trade_counter} {pair} @ ${signal.price:.4f} "
                         f"| RSI={signal.rsi:.1f} | Stop=${signal.stop_price:.4f} "
                         f"| Target=${signal.target_price:.4f} | Fee=${entry_fee:.2f}")

        # Save position
        if 'positions' not in self.state:
            self.state['positions'] = {}

        self.state['positions'][pair] = {
            'pair': pair,
            'entry_price': signal.price,
            'volume': volume,
            'stop_price': signal.stop_price,
            'highest_price': signal.price,
            'entry_time': int(time.time()),
            'target_price': signal.target_price,
            'rsi': signal.rsi,
            'trade_nr': self.trade_counter,
        }

        self.state['total_trades'] = self.trade_counter

        # Log BUY to Excel
        log_buy_to_excel({
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'pair': pair,
            'entry_price': signal.price,
            'volume': volume,
            'amount_usd': self.capital_per_trade,
            'stop_price': signal.stop_price,
            'target_price': signal.target_price,
            'rsi': signal.rsi,
        }, self.trade_counter)

        save_state(self.state)

    def _paper_sell(self, pair: str, signal: Signal, position: Position):
        """Simuleer een sell."""
        pos_data = self.state['positions'].get(pair, {})
        entry_price = pos_data.get('entry_price', position.entry_price)
        volume = pos_data.get('volume', position.volume)
        entry_time = pos_data.get('entry_time', position.entry_time)
        trade_nr = pos_data.get('trade_nr', self.trade_counter)

        # P&L berekening met fees
        gross_pnl = (signal.price - entry_price) * volume
        entry_fee = entry_price * volume * KRAKEN_FEE
        exit_fee = signal.price * volume * KRAKEN_FEE
        total_fee = entry_fee + exit_fee
        net_pnl = gross_pnl - total_fee
        pnl_pct = (signal.price - entry_price) / entry_price * 100

        duration_hours = (time.time() - entry_time) / 3600

        exit_type = "TARGET" if signal.action == 'SELL_TARGET' else "STOP"
        emoji = "WIN" if net_pnl >= 0 else "LOSS"

        self.logger.info(f"  SELL #{trade_nr} {pair} @ ${signal.price:.4f} [{exit_type}] "
                         f"| P&L=${net_pnl:+.2f} ({pnl_pct:+.1f}%) | Fee=${total_fee:.2f} "
                         f"| {emoji} | Duur={duration_hours:.1f}h")

        # Update state
        self.state['total_pnl'] = self.state.get('total_pnl', 0) + net_pnl
        self.state['closed_trades'] = self.state.get('closed_trades', 0) + 1

        if net_pnl >= 0:
            self.state['wins'] = self.state.get('wins', 0) + 1
            self.state['gross_wins'] = self.state.get('gross_wins', 0) + net_pnl
            if net_pnl > self.state.get('biggest_win', 0):
                self.state['biggest_win'] = net_pnl
        else:
            self.state['losses'] = self.state.get('losses', 0) + 1
            self.state['gross_losses'] = self.state.get('gross_losses', 0) + net_pnl
            if net_pnl < self.state.get('biggest_loss', 0):
                self.state['biggest_loss'] = net_pnl

        # Remove position
        del self.state['positions'][pair]

        # Determine exit reason from signal
        exit_reason = signal.reason.split(':')[0] if ':' in signal.reason else signal.reason
        if len(exit_reason) > 40:
            exit_reason = exit_reason[:40]

        # Log completed trade to Excel
        log_trade_to_excel({
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'type': f'SELL ({exit_type})',
            'pair': pair,
            'entry_price': entry_price,
            'exit_price': signal.price,
            'volume': volume,
            'amount_usd': entry_price * volume,
            'total_fee': total_fee,
            'pnl': net_pnl,
            'pnl_pct': pnl_pct,
            'exit_reason': exit_reason,
            'rsi': signal.rsi,
            'duration_hours': duration_hours,
        }, trade_nr, self.state.get('total_pnl', 0))

        save_state(self.state)

    def print_summary(self):
        """Print eindrapport."""
        state = self.state
        total = state.get('closed_trades', 0)
        wins = state.get('wins', 0)
        losses = state.get('losses', 0)
        pnl = state.get('total_pnl', 0)
        biggest_win = state.get('biggest_win', 0)
        biggest_loss = state.get('biggest_loss', 0)
        gross_wins = state.get('gross_wins', 0)
        gross_losses = abs(state.get('gross_losses', 0))
        open_pos = len(state.get('positions', {}))

        wr = (wins / total * 100) if total > 0 else 0
        avg_pnl = (pnl / total) if total > 0 else 0
        pf = (gross_wins / gross_losses) if gross_losses > 0 else 0

        print(f"\n{'='*60}")
        print(f"  PAPER TRADING RAPPORT")
        print(f"{'='*60}")
        print(f"  Strategie:       Dual Confirm V4 OPTIMIZED")
        print(f"  Config:          {self.max_positions}x ${self.capital_per_trade:.0f}")
        print(f"  V4 Features:     VolSpike>2.0x | VolConfirm | BE+3% | RSI<40 | ATR2.0 | TimeMax10")
        print(f"  Kraken fee:      {KRAKEN_FEE*100:.2f}%")
        print(f"  Checks uitgevoerd: {state.get('checks', 0)}")
        print(f"{'='*60}")
        print(f"  Afgesloten trades: {total}")
        print(f"  Winstgevend:       {wins}")
        print(f"  Verliesgevend:     {losses}")
        print(f"  Win rate:          {wr:.1f}%")
        print(f"{'='*60}")
        print(f"  Totaal P&L:        ${pnl:+.2f}")
        print(f"  Gem. P&L/trade:    ${avg_pnl:+.2f}")
        print(f"  Grootste winst:    ${biggest_win:+.2f}")
        print(f"  Grootste verlies:  ${biggest_loss:+.2f}")
        print(f"  Profit Factor:     {pf:.2f}")
        print(f"{'='*60}")
        if open_pos > 0:
            print(f"  Open posities:     {open_pos}")
            for pair, pos in state.get('positions', {}).items():
                print(f"    - {pair}: entry=${pos['entry_price']:.4f}")
        print(f"{'='*60}")
        print(f"  Excel bestand:     {EXCEL_FILE}")
        print(f"  Log bestand:       {LOG_FILE}")
        print(f"{'='*60}\n")

    def run(self):
        """Start paper trading loop."""
        start_time = datetime.now(timezone.utc)
        end_time = start_time + timedelta(hours=self.duration_hours)

        self.logger.info(f"Paper trading gestart tot {end_time.strftime('%Y-%m-%d %H:%M')} UTC")
        self.logger.info(f"Excel output: {EXCEL_FILE}")

        # Reset state voor nieuwe run
        self.state = {
            'positions': {},
            'start_time': start_time.isoformat(),
            'config_capital': self.capital_per_trade,
            'config_max_pos': self.max_positions,
            'total_trades': 0,
            'closed_trades': 0,
            'total_pnl': 0.0,
            'wins': 0,
            'losses': 0,
            'gross_wins': 0.0,
            'gross_losses': 0.0,
            'biggest_win': 0.0,
            'biggest_loss': 0.0,
            'checks': 0,
        }
        self.trade_counter = 0
        save_state(self.state)

        # Maak nieuw Excel
        create_excel()

        # Test verbinding
        if not self.client.test_connection():
            self.logger.error("Kraken verbinding gefaald!")
            return

        # Eerste check meteen
        self.check_all_coins()

        while datetime.now(timezone.utc) < end_time:
            # Bereken volgende 4H candle close
            now = datetime.now(timezone.utc)
            next_4h = now.replace(minute=2, second=0, microsecond=0)
            next_4h_hour = ((now.hour // 4) + 1) * 4
            if next_4h_hour >= 24:
                next_4h = next_4h.replace(hour=0) + timedelta(days=1)
            else:
                next_4h = next_4h.replace(hour=next_4h_hour)

            wait_seconds = (next_4h - now).total_seconds()
            wait_seconds = max(wait_seconds, 60)  # Minimaal 1 min wachten

            remaining = (end_time - now).total_seconds() / 3600
            self.logger.info(f"\nVolgende check: {next_4h.strftime('%H:%M')} UTC "
                             f"(over {wait_seconds/60:.0f} min) | "
                             f"Resterend: {remaining:.1f} uur")

            # Check of we nog door moeten
            if wait_seconds > (end_time - now).total_seconds():
                self.logger.info("Eindtijd bereikt tijdens wachten.")
                break

            try:
                time.sleep(wait_seconds)
            except KeyboardInterrupt:
                self.logger.info("\nPaper trading gestopt door gebruiker (Ctrl+C)")
                break

            # Nieuwe check
            if datetime.now(timezone.utc) < end_time:
                self.check_all_coins()

        # Eindrapport
        self.logger.info("\nPaper trading periode afgelopen!")
        self.print_summary()

        # Update Excel met eindtijd
        update_overview_sheet(self.state, self.state.get('positions', {}))
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Overzicht"]
        ws.cell(row=4, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M') + ' (afgelopen)')
        wb.save(EXCEL_FILE)


def show_report(tag=None):
    """Toon het laatste rapport."""
    global STATE_FILE, EXCEL_FILE
    if tag:
        STATE_FILE = BASE_DIR / f'paper_state_{tag}.json'
        EXCEL_FILE = BASE_DIR / f'paper_trades_{tag}.xlsx'

    if not STATE_FILE.exists():
        print(f"Geen paper trade state gevonden ({STATE_FILE}). Start eerst een paper trade.")
        return

    with open(STATE_FILE, 'r') as f:
        state = json.load(f)

    # Create a temporary PaperTrader just for the summary
    trader = PaperTrader.__new__(PaperTrader)
    trader.state = state
    trader.capital_per_trade = state.get('config_capital', float(os.getenv('CAPITAL_PER_TRADE', '1000')))
    trader.max_positions = state.get('config_max_pos', int(os.getenv('MAX_OPEN_POSITIONS', '2')))
    trader.print_summary()


def show_comparison():
    """Vergelijk beide varianten naast elkaar."""
    variants = []
    for tag in ['2x1000', '3x700']:
        sf = BASE_DIR / f'paper_state_{tag}.json'
        if sf.exists():
            with open(sf, 'r') as f:
                state = json.load(f)
                state['tag'] = tag
                variants.append(state)

    if len(variants) < 2:
        print("Niet genoeg data voor vergelijking. Beide varianten moeten actief zijn.")
        for v in variants:
            tag = v['tag']
            print(f"  {tag}: {v.get('closed_trades', 0)} trades, P&L=${v.get('total_pnl', 0):.2f}")
        return

    print(f"\n{'='*65}")
    print(f"  PAPER TRADING VERGELIJKING")
    print(f"{'='*65}")
    print(f"  {'Metric':<25} {'2x$1000':>15} {'3x$700':>15}")
    print(f"  {'-'*55}")

    for v in variants:
        v['wr'] = (v.get('wins', 0) / v.get('closed_trades', 1) * 100) if v.get('closed_trades', 0) > 0 else 0
        v['avg'] = (v.get('total_pnl', 0) / v.get('closed_trades', 1)) if v.get('closed_trades', 0) > 0 else 0
        gw = v.get('gross_wins', 0)
        gl = abs(v.get('gross_losses', 0))
        v['pf'] = (gw / gl) if gl > 0 else 0

    v1, v2 = variants[0], variants[1]

    rows = [
        ('Afgesloten trades', f"{v1.get('closed_trades', 0)}", f"{v2.get('closed_trades', 0)}"),
        ('Winstgevend', f"{v1.get('wins', 0)}", f"{v2.get('wins', 0)}"),
        ('Verliesgevend', f"{v1.get('losses', 0)}", f"{v2.get('losses', 0)}"),
        ('Win rate', f"{v1['wr']:.1f}%", f"{v2['wr']:.1f}%"),
        ('', '', ''),
        ('Totaal P&L', f"${v1.get('total_pnl', 0):+.2f}", f"${v2.get('total_pnl', 0):+.2f}"),
        ('Gem. P&L/trade', f"${v1['avg']:+.2f}", f"${v2['avg']:+.2f}"),
        ('Grootste winst', f"${v1.get('biggest_win', 0):+.2f}", f"${v2.get('biggest_win', 0):+.2f}"),
        ('Grootste verlies', f"${v1.get('biggest_loss', 0):+.2f}", f"${v2.get('biggest_loss', 0):+.2f}"),
        ('Profit Factor', f"{v1['pf']:.2f}", f"{v2['pf']:.2f}"),
        ('', '', ''),
        ('Open posities', f"{len(v1.get('positions', {}))}", f"{len(v2.get('positions', {}))}"),
        ('Checks', f"{v1.get('checks', 0)}", f"{v2.get('checks', 0)}"),
    ]

    for label, val1, val2 in rows:
        if label:
            print(f"  {label:<25} {val1:>15} {val2:>15}")
        else:
            print(f"  {'-'*55}")

    # Winner
    pnl1 = v1.get('total_pnl', 0)
    pnl2 = v2.get('total_pnl', 0)
    if pnl1 > pnl2:
        print(f"\n  >>> WINNAAR: 2x$1000 (${pnl1-pnl2:+.2f} meer P&L)")
    elif pnl2 > pnl1:
        print(f"\n  >>> WINNAAR: 3x$700 (${pnl2-pnl1:+.2f} meer P&L)")
    else:
        print(f"\n  >>> GELIJK")

    print(f"{'='*65}")
    print(f"  Excel 2x$1000: paper_trades_2x1000.xlsx")
    print(f"  Excel 3x$700:  paper_trades_3x700.xlsx")
    print(f"{'='*65}\n")


# === CLI ===

def main():
    parser = argparse.ArgumentParser(description='Paper Trading Bot met Excel logging')
    parser.add_argument('--hours', type=int, default=48,
                        help='Duur in uren (default: 48)')
    parser.add_argument('--capital', type=float, default=None,
                        help='Capital per trade (override .env)')
    parser.add_argument('--max-positions', type=int, default=None,
                        help='Max open positions (override .env)')
    parser.add_argument('--tag', type=str, default=None,
                        help='Variant tag voor unieke bestanden (bijv. 2x1000, 3x700)')
    parser.add_argument('--report', action='store_true',
                        help='Toon rapport voor een variant')
    parser.add_argument('--compare', action='store_true',
                        help='Vergelijk 2x1000 vs 3x700')
    args = parser.parse_args()

    if args.compare:
        show_comparison()
    elif args.report:
        show_report(tag=args.tag)
    else:
        trader = PaperTrader(
            duration_hours=args.hours,
            capital=args.capital,
            max_positions=args.max_positions,
            tag=args.tag,
        )
        trader.run()


if __name__ == '__main__':
    main()
