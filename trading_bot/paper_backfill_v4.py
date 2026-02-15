#!/usr/bin/env python3
"""
Paper Trade V4 — Backfill + Live (Multi-Exchange)
===================================================
1. Haalt historische 4H candles op van Kraken (backfill)
2. Simuleert V4 strategie vanaf 12 feb 2026 12:00 NL (11:00 UTC)
3. Live fase scant ook Solana DEX coins via CoinScanner + GeckoTerminal
4. Logt alle trades naar Excel met correcte timestamps + bron

Gebruik:
    python paper_backfill_v4.py                # Backfill + live (oneindig)
    python paper_backfill_v4.py --hours 168    # Backfill + 7 dagen live
    python paper_backfill_v4.py --report       # Toon rapport
    python paper_backfill_v4.py --no-solana    # Zonder Solana DEX scanning
"""
import os
import sys
import json
import time
import logging
import argparse
from datetime import datetime, timedelta, timezone
from typing import Optional, List
from pathlib import Path
from dotenv import load_dotenv

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from kraken_client import KrakenClient
from strategy import DualConfirmStrategy, Position, Signal

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / '.env')

TAG = 'v4_live'
EXCEL_FILE = BASE_DIR / f'paper_trades_{TAG}.xlsx'
STATE_FILE = BASE_DIR / f'paper_state_{TAG}.json'
LOG_DIR = BASE_DIR / 'logs'
LOG_DIR.mkdir(exist_ok=True)
LOG_FILE = LOG_DIR / f"paper_{TAG}_{datetime.now().strftime('%Y%m%d_%H%M')}.log"

KRAKEN_FEE = 0.0026
SOLANA_FEE = 0.003  # ~0.3% DEX fee
CAPITAL = 2000.0
MAX_POSITIONS = 1

# Backfill start: 12 feb 2026, 12:00 NL = 11:00 UTC
# Eerste 4H candle close daarna = 12:00 UTC
BACKFILL_START_UTC = datetime(2026, 2, 12, 12, 0, tzinfo=timezone.utc)


def setup_logging():
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s [%(levelname)s] %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S',
        handlers=[
            logging.FileHandler(LOG_FILE),
            logging.StreamHandler()
        ]
    )
    return logging.getLogger('paper_v4')


# ============================================================
# STATE
# ============================================================
def load_state():
    if STATE_FILE.exists():
        with open(STATE_FILE, 'r') as f:
            return json.load(f)
    return new_state()


def new_state():
    return {
        'positions': {},
        'start_time': datetime.now(timezone.utc).isoformat(),
        'total_trades': 0,
        'closed_trades': 0,
        'total_pnl': 0.0,
        'wins': 0,
        'losses': 0,
        'gross_wins': 0.0,
        'gross_losses': 0.0,
        'biggest_win': 0.0,
        'biggest_loss': 0.0,
        'checks': 0,
        'backfill_done': False,
    }


def save_state(state):
    with open(STATE_FILE, 'w') as f:
        json.dump(state, f, indent=2)


# ============================================================
# EXCEL
# ============================================================
def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.title = "Trades"

    headers = [
        'Nr', 'Datum/Tijd', 'Type', 'Coin', 'Bron', 'Entry Prijs', 'Exit Prijs',
        'Volume', 'Bedrag ($)', 'Fee ($)', 'P&L ($)', 'P&L (%)',
        'Exit Reden', 'RSI', 'Duur (uren)', 'Cumulatief P&L ($)'
    ]

    hf = Font(bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = tb

    widths = [5, 18, 6, 12, 10, 12, 12, 14, 12, 8, 10, 9, 25, 6, 10, 14]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # Overzicht sheet
    ws2 = wb.create_sheet("Overzicht")
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 18

    items = [
        ('PAPER TRADING V4 RAPPORT', ''),
        ('', ''),
        ('Start', BACKFILL_START_UTC.strftime('%Y-%m-%d %H:%M UTC')),
        ('Einde', '(lopend)'),
        ('Strategie', 'Dual Confirm V4 OPTIMIZED'),
        ('Budget', f'${CAPITAL:.0f}'),
        ('', ''),
        ('RESULTATEN', ''),
        ('Totaal trades', 0),
        ('Winstgevend', 0),
        ('Verliesgevend', 0),
        ('Win rate', '0%'),
        ('', ''),
        ('Totaal P&L', '$0.00'),
        ('Gemiddelde P&L per trade', '$0.00'),
        ('Grootste winst', '$0.00'),
        ('Grootste verlies', '$0.00'),
        ('Profit Factor', 'N/A'),
        ('', ''),
        ('CONFIGURATIE V4', ''),
        ('Max posities', str(MAX_POSITIONS)),
        ('Capital per trade', f'${CAPITAL:.0f}'),
        ('Kraken fee', '0.26% per side'),
        ('RSI DC/BB max', '40'),
        ('ATR stop mult', '2.0'),
        ('Cooldown', '4 bars (8 na stop)'),
        ('Max stop loss', '15%'),
        ('Volume spike filter', '>2.0x avg'),
        ('Break-even stop', 'Na +3% winst'),
        ('Time max', '10 bars / 40h (V4)'),
        ('Vol confirm', '>1.0x prev bar (V4)'),
        ('Volume min', '50% avg'),
    ]

    tf = Font(bold=True, size=14, color='2F5496')
    sf = Font(bold=True, size=11, color='2F5496')
    nf = Font(size=11)

    for row, (k, v) in enumerate(items, 1):
        ca = ws2.cell(row=row, column=1, value=k)
        cb = ws2.cell(row=row, column=2, value=v)
        if row == 1:
            ca.font = tf
        elif k in ('RESULTATEN', 'CONFIGURATIE V4'):
            ca.font = sf
        else:
            ca.font = nf
            cb.font = nf

    # Open Posities sheet
    ws3 = wb.create_sheet("Open Posities")
    pos_h = ['Coin', 'Entry Prijs', 'Entry Tijd', 'Volume', 'Bedrag ($)',
             'Stop Prijs', 'Huidige Prijs', 'P&L ($)']
    for col, h in enumerate(pos_h, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = hf
        cell.fill = hfill
        cell.alignment = ha
        cell.border = tb
    ws3.freeze_panes = 'A2'

    wb.save(EXCEL_FILE)


def log_trade_excel(trade, trade_nr, cum_pnl):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Trades"]
    row = ws.max_row + 1
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    pnl = trade.get('pnl', 0)
    fill = PatternFill(start_color='E2EFDA' if pnl >= 0 else 'FCE4EC',
                       end_color='E2EFDA' if pnl >= 0 else 'FCE4EC', fill_type='solid')

    vals = [
        trade_nr, trade.get('timestamp', ''), trade.get('type', ''),
        trade.get('pair', '').replace('/USD', ''),
        trade.get('source', 'kraken'),
        trade.get('entry_price', 0), trade.get('exit_price', 0),
        trade.get('volume', 0), trade.get('amount_usd', 0),
        trade.get('total_fee', 0), pnl, trade.get('pnl_pct', 0),
        trade.get('exit_reason', ''), trade.get('rsi', 0),
        trade.get('duration_hours', 0), cum_pnl,
    ]

    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = tb
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
        if col in (6, 7):
            cell.number_format = '#,##0.0000'
        elif col == 8:
            cell.number_format = '#,##0.00000000'
        elif col in (9, 10, 11, 16):
            cell.number_format = '$#,##0.00'
        elif col == 12:
            cell.number_format = '0.00"%"'
        elif col == 15:
            cell.number_format = '0.0'

    wb.save(EXCEL_FILE)


def log_buy_excel(trade, trade_nr):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Trades"]
    row = ws.max_row + 1
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))
    blue = PatternFill(start_color='DAEEF3', end_color='DAEEF3', fill_type='solid')

    fee_pct = trade.get('fee_pct', KRAKEN_FEE)
    fee = trade.get('amount_usd', 0) * fee_pct
    vals = [
        trade_nr, trade.get('timestamp', ''), 'BUY',
        trade.get('pair', '').replace('/USD', ''),
        trade.get('source', 'kraken'),
        trade.get('entry_price', 0), '',
        trade.get('volume', 0), trade.get('amount_usd', 0),
        fee, '', '',
        f"Stop: ${trade.get('stop_price', 0):.4f}",
        trade.get('rsi', 0), '', '',
    ]

    for col, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = tb
        cell.fill = blue
        cell.alignment = Alignment(horizontal='center')
        if col == 6 and isinstance(val, (int, float)):
            cell.number_format = '#,##0.0000'
        elif col == 8 and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00000000'
        elif col in (9, 10) and isinstance(val, (int, float)):
            cell.number_format = '$#,##0.00'

    wb.save(EXCEL_FILE)


def update_overview(state):
    if not EXCEL_FILE.exists():
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Overzicht"]

    total = state.get('closed_trades', 0)
    wins = state.get('wins', 0)
    losses = state.get('losses', 0)
    pnl = state.get('total_pnl', 0.0)
    gw = state.get('gross_wins', 0.0)
    gl = abs(state.get('gross_losses', 0.0))

    wr = (wins / total * 100) if total > 0 else 0
    avg = (pnl / total) if total > 0 else 0
    pf = (gw / gl) if gl > 0 else 'N/A'

    ws.cell(row=4, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M') + ' (lopend)')
    ws.cell(row=9, column=2, value=total)
    ws.cell(row=10, column=2, value=wins)
    ws.cell(row=11, column=2, value=losses)
    ws.cell(row=12, column=2, value=f'{wr:.1f}%')
    ws.cell(row=14, column=2, value=f'${pnl:.2f}')
    ws.cell(row=15, column=2, value=f'${avg:.2f}')
    ws.cell(row=16, column=2, value=f'${state.get("biggest_win", 0):.2f}')
    ws.cell(row=17, column=2, value=f'${state.get("biggest_loss", 0):.2f}')
    ws.cell(row=18, column=2, value=f'{pf:.2f}' if isinstance(pf, float) else pf)

    wb.save(EXCEL_FILE)


# ============================================================
# V4 STRATEGY FACTORY
# ============================================================
def make_v4_strategy():
    return DualConfirmStrategy(
        rsi_dc_max=40, rsi_bb_max=40, rsi_sell=70,
        atr_stop_mult=2.0, cooldown_bars=4,
        cooldown_after_stop=8,
        max_stop_loss_pct=15.0,
        volume_spike_filter=True, volume_spike_mult=2.0,
        breakeven_stop=True, breakeven_trigger_pct=3.0,
        volume_min_pct=0.5,
        time_max_bars=10,
        vol_confirm=True, vol_confirm_mult=1.0,
    )


# ============================================================
# BACKFILL ENGINE
# ============================================================
def run_backfill(client, logger, enable_solana=True):
    """
    Backfill paper trades van 12 feb 12:00 UTC tot nu.
    Haalt historische candles op van Kraken + Solana DEX (GeckoTerminal).
    """
    logger.info("=" * 60)
    logger.info("  BACKFILL FASE — V4 Paper Trade (Multi-Exchange)")
    logger.info(f"  Vanaf: {BACKFILL_START_UTC.strftime('%Y-%m-%d %H:%M UTC')}")
    logger.info(f"  Tot:   {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    logger.info("=" * 60)

    # Dynamische coin discovery: haal ALLE Kraken USD pairs op
    coins_str = os.getenv('COINS', '')
    if coins_str:
        coins = [c.strip() for c in coins_str.split(',') if c.strip()]
        logger.info(f"Kraken coins (uit .env): {len(coins)}")
    else:
        coins = client.get_all_tradeable_pairs()
        logger.info(f"Kraken coins (dynamisch ontdekt): {len(coins)}")

    # Per-coin strategy instances
    strategies = {}
    for pair in coins:
        strategies[pair] = make_v4_strategy()

    # Coin metadata: welke bron + fee per pair
    coin_meta = {}  # pair → {'source': str, 'fee_pct': float}
    for pair in coins:
        coin_meta[pair] = {'source': 'kraken', 'fee_pct': KRAKEN_FEE}

    state = new_state()
    state['backfill_done'] = False
    trade_counter = 0

    # ─── 1. Kraken candles ophalen (720 bars = 120 dagen) ───
    logger.info("Candles ophalen van Kraken...")
    all_candles = {}
    errors = 0
    for i, pair in enumerate(coins):
        try:
            candles = client.get_ohlc(pair, interval=240)
            if candles and len(candles) > 30:
                all_candles[pair] = candles
            if (i + 1) % 50 == 0:
                logger.info(f"  {i+1}/{len(coins)} coins geladen...")
                time.sleep(1)  # Rate limiting
        except Exception as e:
            errors += 1
            if errors <= 10:
                logger.warning(f"  Skip {pair}: {e}")
        # Kraken rate limit: ~1 req/sec for public
        time.sleep(0.35)

    logger.info(f"Kraken candles geladen: {len(all_candles)} coins ({errors} errors)")

    # ─── 2. Solana DEX candles ophalen (GeckoTerminal, 180 dagen) ───
    solana_coins = []
    if enable_solana:
        logger.info("Solana token discovery voor backfill...")
        solana_coins = _scan_solana_coins(logger)

        # Dedup: skip coins die al op Kraken staan
        kraken_symbols = {c.split('/')[0] for c in coins}
        solana_coins = [
            sc for sc in solana_coins
            if sc['symbol'] not in kraken_symbols
        ]
        logger.info(f"Solana: {len(solana_coins)} unieke coins (na Kraken dedup)")

        gecko = _get_gecko()
        for sc in solana_coins:
            pair = sc['pair']
            address = sc.get('address', '')
            if not address:
                continue
            try:
                candles = gecko.get_ohlc('solana', address, interval=240)
                if candles and len(candles) > 30:
                    # GeckoTerminal gebruikt 'time' als key, normaliseer naar 'timestamp'
                    for c in candles:
                        if 'time' in c and 'timestamp' not in c:
                            c['timestamp'] = c['time']
                    all_candles[pair] = candles
                    strategies[pair] = make_v4_strategy()
                    coin_meta[pair] = {'source': 'dex_solana', 'fee_pct': SOLANA_FEE}
                    logger.info(f"  Solana {sc['symbol']}: {len(candles)} candles geladen")
                else:
                    logger.info(f"  Solana {sc['symbol']}: onvoldoende candles, overgeslagen")
            except Exception as e:
                logger.warning(f"  Solana {sc['symbol']} error: {e}")

        logger.info(f"Totaal candles geladen: {len(all_candles)} coins "
                     f"(Kraken: {len(coins)}, Solana: {len([p for p in all_candles if coin_meta.get(p, {}).get('source') == 'dex_solana'])})")

    # Bepaal welke 4H bars we moeten simuleren
    # 4H candle closes: 00, 04, 08, 12, 16, 20 UTC
    now_utc = datetime.now(timezone.utc)
    check_times = []
    t = BACKFILL_START_UTC
    while t <= now_utc:
        check_times.append(t)
        t += timedelta(hours=4)

    # Verwijder de huidige/toekomstige candle (nog niet gesloten)
    # De huidige 4H candle is nog bezig, dus verwijder die
    current_4h = now_utc.replace(minute=0, second=0, microsecond=0)
    current_4h_hour = (now_utc.hour // 4) * 4
    current_4h = current_4h.replace(hour=current_4h_hour)
    check_times = [t for t in check_times if t < current_4h]

    logger.info(f"Te simuleren checks: {len(check_times)}")
    for ct in check_times:
        logger.info(f"  {ct.strftime('%Y-%m-%d %H:%M UTC')}")

    # Simuleer elke check
    for check_time in check_times:
        state['checks'] += 1
        check_ts = int(check_time.timestamp())
        check_str = check_time.strftime('%Y-%m-%d %H:%M')

        logger.info(f"\n{'='*50}")
        logger.info(f"BACKFILL CHECK #{state['checks']} — {check_str} UTC")
        logger.info(f"{'='*50}")

        for pair in all_candles:
            candles = all_candles[pair]

            # Vind de candle die overeenkomt met deze check_time
            # Candle timestamp is de OPEN time, we zoeken close time
            # close_time = open_time + 4h
            # Dus we zoeken candles met open_time <= check_ts - 14400
            target_open_ts = check_ts - 14400  # 4h = 14400 sec

            # Filter candles tot en met deze check_time
            # Kraken candle timestamp = open time
            valid_candles = []
            for c in candles:
                c_ts = c.get('timestamp', 0)
                # Candle is geldig als: open_time + 14400 <= check_ts
                # (de candle moet gesloten zijn op check_time)
                if c_ts + 14400 <= check_ts:
                    valid_candles.append(c)
                elif c_ts <= check_ts:
                    # Huidige candle (nog niet gesloten) - neem mee als laatste
                    valid_candles.append(c)

            if len(valid_candles) < 30:
                continue

            # Check positie
            pos_data = state['positions'].get(pair)
            position = None
            if pos_data:
                position = Position(
                    pair=pos_data['pair'],
                    entry_price=pos_data['entry_price'],
                    volume=pos_data['volume'],
                    stop_price=pos_data['stop_price'],
                    highest_price=pos_data['highest_price'],
                    entry_time=pos_data['entry_time'],
                )

            strategy = strategies[pair]
            signal = strategy.analyze(valid_candles, position, pair)

            # Coin-specifieke metadata
            meta = coin_meta.get(pair, {'source': 'kraken', 'fee_pct': KRAKEN_FEE})
            source = meta['source']
            fee_pct = meta['fee_pct']
            source_label = 'SOL' if source.startswith('dex') else 'KRK'

            # SELL
            if signal.action in ('SELL_TARGET', 'SELL_STOP') and position:
                entry_price = pos_data['entry_price']
                volume = pos_data['volume']
                entry_time = pos_data['entry_time']
                trade_nr = pos_data.get('trade_nr', trade_counter)
                pos_fee = pos_data.get('fee_pct', fee_pct)

                gross_pnl = (signal.price - entry_price) * volume
                entry_fee = entry_price * volume * pos_fee
                exit_fee = signal.price * volume * pos_fee
                total_fee = entry_fee + exit_fee
                net_pnl = gross_pnl - total_fee
                pnl_pct = (signal.price - entry_price) / entry_price * 100
                duration_h = (check_ts - entry_time) / 3600

                exit_type = "TARGET" if signal.action == 'SELL_TARGET' else "STOP"
                reason = signal.reason.split(':')[0] if ':' in signal.reason else signal.reason

                logger.info(f"  SELL #{trade_nr} [{source_label}] {pair} @ ${signal.price:.4f} [{exit_type}] "
                           f"| P&L=${net_pnl:+.2f} ({pnl_pct:+.1f}%) | {reason}")

                state['total_pnl'] += net_pnl
                state['closed_trades'] += 1
                if net_pnl >= 0:
                    state['wins'] += 1
                    state['gross_wins'] += net_pnl
                    if net_pnl > state['biggest_win']:
                        state['biggest_win'] = net_pnl
                else:
                    state['losses'] += 1
                    state['gross_losses'] += net_pnl
                    if net_pnl < state['biggest_loss']:
                        state['biggest_loss'] = net_pnl

                del state['positions'][pair]

                log_trade_excel({
                    'timestamp': check_str,
                    'type': f'SELL ({exit_type})',
                    'pair': pair,
                    'source': source,
                    'entry_price': entry_price,
                    'exit_price': signal.price,
                    'volume': volume,
                    'amount_usd': entry_price * volume,
                    'total_fee': total_fee,
                    'pnl': net_pnl,
                    'pnl_pct': pnl_pct,
                    'exit_reason': reason,
                    'rsi': signal.rsi,
                    'duration_hours': duration_h,
                }, trade_nr, state['total_pnl'])

            # BUY
            elif signal.action == 'BUY' and pair not in state.get('positions', {}):
                num_open = len(state.get('positions', {}))
                if num_open >= MAX_POSITIONS:
                    continue

                trade_counter += 1
                volume = CAPITAL / signal.price

                logger.info(f"  BUY #{trade_counter} [{source_label}] {pair} @ ${signal.price:.4f} "
                           f"| RSI={signal.rsi:.1f} | Stop=${signal.stop_price:.4f}")

                state['positions'][pair] = {
                    'pair': pair,
                    'entry_price': signal.price,
                    'volume': volume,
                    'stop_price': signal.stop_price,
                    'highest_price': signal.price,
                    'entry_time': check_ts,
                    'trade_nr': trade_counter,
                    'rsi': signal.rsi,
                    'source': source,
                    'fee_pct': fee_pct,
                }
                state['total_trades'] = trade_counter

                log_buy_excel({
                    'timestamp': check_str,
                    'pair': pair,
                    'source': source,
                    'entry_price': signal.price,
                    'volume': volume,
                    'amount_usd': CAPITAL,
                    'stop_price': signal.stop_price,
                    'rsi': signal.rsi,
                    'fee_pct': fee_pct,
                }, trade_counter)

            # HOLD — update tracking
            elif signal.action == 'HOLD' and position and pair in state['positions']:
                current_price = valid_candles[-1]['close']
                if current_price > state['positions'][pair]['highest_price']:
                    state['positions'][pair]['highest_price'] = current_price
                if position.stop_price > state['positions'][pair]['stop_price']:
                    state['positions'][pair]['stop_price'] = position.stop_price

        save_state(state)

    state['backfill_done'] = True
    save_state(state)
    update_overview(state)

    sol_count = len([p for p in all_candles if coin_meta.get(p, {}).get('source') == 'dex_solana'])
    logger.info(f"\n{'='*60}")
    logger.info(f"  BACKFILL COMPLEET (Multi-Exchange)")
    logger.info(f"  Bronnen: Kraken ({len(all_candles) - sol_count}) + Solana ({sol_count})")
    logger.info(f"  Checks: {state['checks']}")
    logger.info(f"  Trades: {state['closed_trades']} gesloten, {len(state['positions'])} open")
    logger.info(f"  P&L: ${state['total_pnl']:+.2f}")
    logger.info(f"{'='*60}")

    return state, trade_counter, strategies


# ============================================================
# LIVE PAPER TRADING (continuation after backfill)
# ============================================================
def _get_fee_for_source(source: str) -> float:
    """Return fee percentage voor een bron."""
    if source.startswith('dex'):
        return SOLANA_FEE
    return KRAKEN_FEE


def _scan_solana_coins(logger) -> List[dict]:
    """
    Scan Solana DEX coins via CoinScanner.
    Cached per sessie (elke 12 uur hernieuwd).
    """
    try:
        from coin_scanner import CoinScanner
        scanner = CoinScanner(
            enable_kraken=False,  # Niet nodig, Kraken coins uit .env
            enable_mexc=False,
            enable_dex=True,
            enable_halal=True,
            dex_chains=['solana'],
        )
        solana_coins = scanner.scan_solana_auto()
        logger.info(f"Solana scan: {len(solana_coins)} veilige tokens gevonden")
        return solana_coins
    except Exception as e:
        logger.warning(f"Solana scan gefaald: {e}")
        return []


_gecko_client = None

def _get_gecko():
    """Singleton GeckoTerminal client (rate limiter behouden over calls)."""
    global _gecko_client
    if _gecko_client is None:
        from dex_manager import GeckoTerminalClient
        _gecko_client = GeckoTerminalClient()
    return _gecko_client


def _get_solana_candles(coin_info: dict, logger) -> Optional[list]:
    """Haal 4H candles op voor een Solana coin via GeckoTerminal."""
    try:
        gecko = _get_gecko()
        address = coin_info.get('address', '')
        if not address:
            return None
        candles = gecko.get_ohlc('solana', address, interval=240)
        if candles and len(candles) >= 30:
            return candles
        return None
    except Exception as e:
        logger.warning(f"GeckoTerminal error {coin_info.get('symbol', '?')}: {e}")
        return None


def run_live(client, logger, state, trade_counter, strategies,
             duration_hours=None, enable_solana=True):
    """Continue paper trading live na backfill."""
    # Dynamische coin discovery: haal ALLE Kraken USD pairs op
    coins_str = os.getenv('COINS', '')
    if coins_str:
        coins = [c.strip() for c in coins_str.split(',') if c.strip()]
    else:
        coins = client.get_all_tradeable_pairs()

    if duration_hours:
        end_time = datetime.now(timezone.utc) + timedelta(hours=duration_hours)
        logger.info(f"\nLive paper trading tot {end_time.strftime('%Y-%m-%d %H:%M UTC')} ({duration_hours}h)")
    else:
        end_time = None
        logger.info(f"\nLive paper trading — oneindig (Ctrl+C om te stoppen)")

    if enable_solana:
        logger.info("Solana DEX scanning INGESCHAKELD")

    # Solana coin cache (hernieuwd elke 12 uur)
    solana_coins = []
    solana_last_scan = 0
    SOLANA_SCAN_INTERVAL = 12 * 3600  # 12 uur

    while True:
        if end_time and datetime.now(timezone.utc) >= end_time:
            logger.info("Eindtijd bereikt!")
            break

        # Bereken volgende 4H candle close + 2 min
        now = datetime.now(timezone.utc)
        next_4h = now.replace(minute=2, second=0, microsecond=0)
        next_4h_hour = ((now.hour // 4) + 1) * 4
        if next_4h_hour >= 24:
            next_4h = next_4h.replace(hour=0) + timedelta(days=1)
        else:
            next_4h = next_4h.replace(hour=next_4h_hour)

        wait_seconds = (next_4h - now).total_seconds()
        wait_seconds = max(wait_seconds, 60)

        sol_count = len(solana_coins) if enable_solana else 0
        if end_time:
            remaining = (end_time - now).total_seconds() / 3600
            logger.info(f"\nVolgende check: {next_4h.strftime('%H:%M UTC')} "
                        f"(over {wait_seconds/60:.0f} min) | Resterend: {remaining:.1f}h"
                        f" | Kraken: {len(coins)}, Solana: {sol_count}")
        else:
            logger.info(f"\nVolgende check: {next_4h.strftime('%H:%M UTC')} "
                        f"(over {wait_seconds/60:.0f} min)"
                        f" | Kraken: {len(coins)}, Solana: {sol_count}")

        try:
            time.sleep(wait_seconds)
        except KeyboardInterrupt:
            logger.info("\nGestopt door gebruiker (Ctrl+C)")
            break

        # ─── Solana coin discovery (elke 12 uur) ───
        if enable_solana and (time.time() - solana_last_scan) > SOLANA_SCAN_INTERVAL:
            logger.info("Solana token discovery starten...")
            solana_coins = _scan_solana_coins(logger)
            solana_last_scan = time.time()
            # Filter coins die al op Kraken staan (dedup)
            kraken_symbols = {c.split('/')[0] for c in coins}
            solana_coins = [
                sc for sc in solana_coins
                if sc['symbol'] not in kraken_symbols
            ]
            logger.info(f"Solana: {len(solana_coins)} unieke coins (na Kraken dedup)")

        # Live check
        state['checks'] += 1
        check_str = datetime.now().strftime('%Y-%m-%d %H:%M')
        logger.info(f"\n{'='*50}")
        logger.info(f"LIVE CHECK #{state['checks']} — {check_str}")
        logger.info(f"Open: {len(state.get('positions', {}))}/{MAX_POSITIONS} | P&L: ${state['total_pnl']:+.2f}")
        logger.info(f"{'='*50}")

        # Shared trade counter ref (mutable list voor pass-by-reference)
        tc_ref = [trade_counter]

        # ─── 1. Kraken coins scannen ───
        logger.info(f"--- Kraken ({len(coins)} coins) ---")
        for pair in coins:
            try:
                candles = client.get_ohlc(pair, interval=240)
                if not candles or len(candles) < 30:
                    continue

                _process_coin(
                    pair=pair,
                    candles=candles,
                    source='kraken',
                    fee_pct=KRAKEN_FEE,
                    state=state,
                    strategies=strategies,
                    trade_counter_ref=tc_ref,
                    check_str=check_str,
                    logger=logger,
                )

            except Exception as e:
                logger.warning(f"  Error {pair}: {e}")

        # ─── 2. Solana DEX coins scannen ───
        if enable_solana and solana_coins:
            logger.info(f"--- Solana DEX ({len(solana_coins)} coins) ---")
            for coin_info in solana_coins:
                try:
                    pair = coin_info['pair']
                    candles = _get_solana_candles(coin_info, logger)
                    if not candles:
                        continue

                    _process_coin(
                        pair=pair,
                        candles=candles,
                        source='dex_solana',
                        fee_pct=SOLANA_FEE,
                        state=state,
                        strategies=strategies,
                        trade_counter_ref=tc_ref,
                        check_str=check_str,
                        logger=logger,
                    )

                except Exception as e:
                    logger.warning(f"  Error Solana {coin_info.get('symbol', '?')}: {e}")

        # Update trade counter van ref
        trade_counter = tc_ref[0]

        save_state(state)
        update_overview(state)

        # Status summary
        logger.info(f"\nCheck compleet | Open: {len(state.get('positions', {}))}/{MAX_POSITIONS} "
                    f"| Trades: {state['closed_trades']} | P&L: ${state['total_pnl']:+.2f}")

    # Final report
    print_summary(state)
    update_overview(state)


def _process_coin(pair, candles, source, fee_pct, state, strategies,
                  trade_counter_ref, check_str, logger):
    """
    Verwerk een coin (Kraken of Solana) voor signalen.
    Gedeelde logica voor buy/sell/hold.
    """
    pos_data = state['positions'].get(pair)
    position = None
    if pos_data:
        position = Position(
            pair=pos_data['pair'],
            entry_price=pos_data['entry_price'],
            volume=pos_data['volume'],
            stop_price=pos_data['stop_price'],
            highest_price=pos_data['highest_price'],
            entry_time=pos_data['entry_time'],
        )

    if pair not in strategies:
        strategies[pair] = make_v4_strategy()
    strategy = strategies[pair]
    signal = strategy.analyze(candles, position, pair)

    check_ts = int(time.time())
    trade_counter = trade_counter_ref[0]
    source_label = 'SOL' if source.startswith('dex') else 'KRK'

    # SELL
    if signal.action in ('SELL_TARGET', 'SELL_STOP') and position:
        entry_price = pos_data['entry_price']
        volume = pos_data['volume']
        entry_time = pos_data['entry_time']
        trade_nr = pos_data.get('trade_nr', trade_counter)
        pos_fee = pos_data.get('fee_pct', fee_pct)

        gross_pnl = (signal.price - entry_price) * volume
        entry_fee = entry_price * volume * pos_fee
        exit_fee = signal.price * volume * pos_fee
        total_fee = entry_fee + exit_fee
        net_pnl = gross_pnl - total_fee
        pnl_pct = (signal.price - entry_price) / entry_price * 100
        duration_h = (check_ts - entry_time) / 3600

        exit_type = "TARGET" if signal.action == 'SELL_TARGET' else "STOP"
        reason = signal.reason.split(':')[0] if ':' in signal.reason else signal.reason

        logger.info(f"  🔴 SELL #{trade_nr} [{source_label}] {pair} @ ${signal.price:.4f} [{exit_type}] "
                   f"| P&L=${net_pnl:+.2f} ({pnl_pct:+.1f}%) | {reason}")

        state['total_pnl'] += net_pnl
        state['closed_trades'] += 1
        if net_pnl >= 0:
            state['wins'] += 1
            state['gross_wins'] += net_pnl
            if net_pnl > state['biggest_win']:
                state['biggest_win'] = net_pnl
        else:
            state['losses'] += 1
            state['gross_losses'] += net_pnl
            if net_pnl < state['biggest_loss']:
                state['biggest_loss'] = net_pnl

        del state['positions'][pair]

        log_trade_excel({
            'timestamp': check_str,
            'type': f'SELL ({exit_type})',
            'pair': pair,
            'source': source,
            'entry_price': entry_price,
            'exit_price': signal.price,
            'volume': volume,
            'amount_usd': entry_price * volume,
            'total_fee': total_fee,
            'pnl': net_pnl,
            'pnl_pct': pnl_pct,
            'exit_reason': reason,
            'rsi': signal.rsi,
            'duration_hours': duration_h,
        }, trade_nr, state['total_pnl'])

    # BUY
    elif signal.action == 'BUY' and pair not in state.get('positions', {}):
        if len(state.get('positions', {})) >= MAX_POSITIONS:
            return

        trade_counter += 1
        volume = CAPITAL / signal.price

        logger.info(f"  🟢 BUY #{trade_counter} [{source_label}] {pair} @ ${signal.price:.4f} "
                   f"| RSI={signal.rsi:.1f} | Stop=${signal.stop_price:.4f}")

        state['positions'][pair] = {
            'pair': pair,
            'entry_price': signal.price,
            'volume': volume,
            'stop_price': signal.stop_price,
            'highest_price': signal.price,
            'entry_time': check_ts,
            'trade_nr': trade_counter,
            'rsi': signal.rsi,
            'source': source,
            'fee_pct': fee_pct,
        }
        state['total_trades'] = trade_counter
        trade_counter_ref[0] = trade_counter

        log_buy_excel({
            'timestamp': check_str,
            'pair': pair,
            'source': source,
            'entry_price': signal.price,
            'volume': volume,
            'amount_usd': CAPITAL,
            'stop_price': signal.stop_price,
            'rsi': signal.rsi,
            'fee_pct': fee_pct,
        }, trade_counter)

    # HOLD
    elif signal.action == 'HOLD' and position and pair in state['positions']:
        current_price = candles[-1]['close']
        if current_price > state['positions'][pair]['highest_price']:
            state['positions'][pair]['highest_price'] = current_price
        if position.stop_price > state['positions'][pair]['stop_price']:
            state['positions'][pair]['stop_price'] = position.stop_price


def print_summary(state):
    total = state.get('closed_trades', 0)
    wins = state.get('wins', 0)
    losses = state.get('losses', 0)
    pnl = state.get('total_pnl', 0)
    gw = state.get('gross_wins', 0)
    gl = abs(state.get('gross_losses', 0))
    wr = (wins / total * 100) if total > 0 else 0
    avg = (pnl / total) if total > 0 else 0
    pf = (gw / gl) if gl > 0 else 0

    print(f"\n{'='*60}")
    print(f"  PAPER TRADING V4 — RAPPORT (Multi-Exchange)")
    print(f"{'='*60}")
    print(f"  Config:          {MAX_POSITIONS}x ${CAPITAL:.0f}")
    print(f"  V4 Features:     VolSpike>2.0x | VolConf | BE+3% | RSI<40 | ATR2.0 | TMB10")
    print(f"  Bronnen:         Kraken (0.26%) + Solana DEX (0.30%)")
    print(f"  Checks:          {state.get('checks', 0)}")
    print(f"{'='*60}")
    print(f"  Afgesloten:      {total}")
    print(f"  Winstgevend:     {wins}")
    print(f"  Verliesgevend:   {losses}")
    print(f"  Win rate:        {wr:.1f}%")
    print(f"{'='*60}")
    print(f"  Totaal P&L:      ${pnl:+.2f}")
    print(f"  Gem. P&L/trade:  ${avg:+.2f}")
    print(f"  Grootste winst:  ${state.get('biggest_win', 0):+.2f}")
    print(f"  Grootste verlies:${state.get('biggest_loss', 0):+.2f}")
    print(f"  Profit Factor:   {pf:.2f}")
    print(f"{'='*60}")
    if state.get('positions'):
        print(f"  Open posities:   {len(state['positions'])}")
        for pair, pos in state['positions'].items():
            print(f"    - {pair}: entry=${pos['entry_price']:.4f}")
    print(f"  Excel: {EXCEL_FILE}")
    print(f"{'='*60}\n")


# ============================================================
# MAIN
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='Paper Trade V4 — Backfill + Live (Multi-Exchange)')
    parser.add_argument('--hours', type=int, default=None,
                        help='Live duur na backfill (default: oneindig)')
    parser.add_argument('--report', action='store_true',
                        help='Toon rapport')
    parser.add_argument('--no-backfill', action='store_true',
                        help='Skip backfill, alleen live')
    parser.add_argument('--no-solana', action='store_true',
                        help='Solana DEX scanning uitschakelen')
    args = parser.parse_args()

    if args.report:
        if STATE_FILE.exists():
            with open(STATE_FILE) as f:
                state = json.load(f)
            print_summary(state)
        else:
            print("Geen state gevonden. Start eerst paper trading.")
        return

    logger = setup_logging()

    # Kraken client
    api_key = os.getenv('KRAKEN_API_KEY', '')
    private_key = os.getenv('KRAKEN_PRIVATE_KEY', '')
    if not api_key or not private_key:
        logger.error("KRAKEN_API_KEY of KRAKEN_PRIVATE_KEY niet ingesteld")
        sys.exit(1)

    client = KrakenClient(api_key, private_key)

    if not client.test_connection():
        logger.error("Kraken verbinding gefaald!")
        sys.exit(1)
    logger.info("Kraken verbinding OK")

    # Fresh start
    create_excel()
    state = new_state()
    trade_counter = 0
    strategies = {}

    enable_solana = not args.no_solana

    # Phase 1: Backfill (incl. Solana via GeckoTerminal)
    if not args.no_backfill:
        state, trade_counter, strategies = run_backfill(client, logger, enable_solana)
    else:
        logger.info("Backfill overgeslagen (--no-backfill)")

    # Phase 2: Live (met optionele Solana DEX scanning)
    run_live(client, logger, state, trade_counter, strategies,
             args.hours, enable_solana=enable_solana)


if __name__ == '__main__':
    main()
